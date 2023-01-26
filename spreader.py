from openpyxl import *
from random import randint
from math import ceil


def create_dict_from_table(values_dict: dict, row: int, start_column_values: int) -> dict:
    """ Формируем словарь из ячеек со значениями отличными от None, словарь (изначально пустой) передаем извне """

    for cell in row[start_column_values:]:
        if cell.value is not None:
            values_dict.setdefault(cell.coordinate, cell.value)
    return values_dict


def find_max_and_values(values_dict: dict, limit_value=300) -> tuple[int, int, int, int, list, int]:
    """ Находим из словаря максимальное количество продаж (и формируем из них список), общее число продаж,
        количество оптовых продаж, сумму оптовых продаж
    """


    sum_sales = 0  # Общее количество проданных товаров
    sum_sales_count = 0  # Количество ячеек с продажами
    opt_sum_sales = 0  # Общее количество товаров, проданных оптом
    opt_sales_count = 0  # Количество ячеек с оптовыми продажами
    opt_sales_list = []  # Список оптовых продаж, добавляется кортеж (ячейка, значение)

    for key, value in values_dict.items():
        if value >= limit_value:
            opt_sales_list.append((key, value))
            opt_sum_sales += value
            values_dict[key] = randint(0, 2)  # Приравниваем значения больше установленного предела к 0...2
            opt_sales_count += 1
        sum_sales += value
        sum_sales_count += 1

    clear_sales = sum_sales - opt_sum_sales

    return sum_sales, sum_sales_count, opt_sum_sales, opt_sales_count, opt_sales_list, clear_sales


def spread_values(values_dict: dict, clear_sales: int, sum_sales: int, opt_sum_sales: int, limit_value=300)\
        -> tuple[dict, int, int]:
    """ Распределяем значения по таблице, в соответствии с их процентным соотношением. Значения больше предельных
        уменьшаем на 1...10
    """

    new_values_dict = values_dict.copy()
    sum_values_dict = sum([value for value in new_values_dict.values()])

    for key, value in new_values_dict.items():
        if value != 1:
            new_value = ceil(new_values_dict[key] + (new_values_dict[key] / clear_sales) * opt_sum_sales)
            difference = new_value - value

            if new_value < limit_value:
                if sum_values_dict + difference < sum_sales:
                    new_values_dict[key] = new_value
                    sum_values_dict += difference
                else:
                    new_diff = sum_sales - sum_values_dict
                    new_values_dict[key] = value + new_diff
                    sum_values_dict += new_diff
                    break
            else:
                corr = randint(1, 10)
                new_value = limit_value - corr
                difference = new_value - value
                new_values_dict[key] = new_value
                sum_values_dict += difference

    # Сумма новых значений в словаре
    new_values_dict_sum = sum([value for value in new_values_dict.values()])

    # Остаток после разбивки
    remain = sum_sales - sum_values_dict

    return new_values_dict, remain, new_values_dict_sum


def spread_remain(new_values_dict: dict, remain: int, limit_value=300) -> tuple[dict, int]:
    """ Распределение остатка (один прогон) """

    remain_dec = remain % 10
    remain_all = (remain // 10) * 10

    for key, value in new_values_dict.items():
        if remain_dec:
            if 31 <= value and (value + remain_dec) < limit_value:
                new_values_dict[key] += remain_dec
                remain_dec = 0
        elif remain_all:
            if remain_all > 25:
                if 51 <= value and (value + 25) < limit_value:
                    new_values_dict[key] += 25
                    remain_all -= 25
            else:
                if 51 <= value and (value + remain_all) < limit_value:
                    new_values_dict[key] += remain_all
                    remain_all -= remain_all
                    break
                else:
                    continue
        else:
            continue

    remain = remain_all

    return new_values_dict, remain


def create_new_file(file_name: str, result_values_list: list[dict]) -> None:
    """ Создаем новый файл и копируем в него данные из получившегося словаря """

    wb = load_workbook(filename=file_name)
    ws = wb.active

    for new_values_dict in result_values_list:
        for key, value in new_values_dict.items():
            ws[key] = value

    wb.save('new_' + file_name)


def create_result_file(result: list[dict]) -> None:
    """ Создаем файл с результатами разбивки для всех товаров в исходном файле """

    wb = Workbook()
    ws = wb.active
    ws.title = 'Результаты'

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

    row_counter = 1

    for item in result:
        for key, value in item.items():
            ws.cell(row=row_counter, column=1).value = key
            ws.cell(row=row_counter, column=2).value = value
            row_counter += 1
        row_counter += 1

    wb.save(filename='Результаты.xlsx')
