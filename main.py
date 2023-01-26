from spreader import *
from app.app import *
from openpyxl.utils.cell import coordinate_to_tuple


def main_func(file_name, start_cell):
    """ Основная рабочая функция, в качестве аргументов передается имя файла и начальная ячейка"""

    book = load_workbook(filename=file_name)
    sheet = book.active

    # Находим координаты первой ячейки и количество строк и столбцов в таблице
    start_cell_x, start_cell_y = coordinate_to_tuple(start_cell)
    start_column_values = start_cell_y + 2
    count_max_column = sheet.max_column
    count_max_row = sheet.max_row

    # Объявляем результирующий список словарей значений и промежуточный словарь
    values_dict = {}  # Промежуточный словарь с распределенными значениями (в пределах одного товара)
    result_values_list = []  # Итоговый список из промежуточных словарей, будет добавляться в конечных файл
    calculated_values_dict = {}  # Промежуточный словарь с рассчитанными значениями (в пределах одного товара)
    result_calculated_values = []  # Итоговый список из промежуточных словарей с рассчитанными значениями

    # Главный цикл
    for row in sheet.iter_rows(min_row=start_cell_x, max_row=count_max_row,
                               min_col=start_cell_y, max_col=count_max_column):

        x, y = coordinate_to_tuple(row[0].coordinate)

        if x == count_max_row:
            values_dict = create_dict_from_table(values_dict, row, start_column_values)

        if type(row[1].value) == str or x == count_max_row:

            if values_dict:
                sum_sales, sum_sales_count, opt_sum_sales, opt_sales_count, opt_sales_list, clear_sales \
                    = find_max_and_values(values_dict)

                new_values_dict, remain, new_values_dict_sum = spread_values(values_dict, clear_sales, sum_sales,
                                                                             opt_sum_sales)

                calculated_values_dict['Продажи тотал, шт.:'] = sum_sales
                calculated_values_dict['Продажи розница, шт.:'] = clear_sales
                calculated_values_dict['Продажи ОПТ, шт.:'] = opt_sum_sales
                calculated_values_dict['Кол-во оптовых продаж:'] = opt_sales_count
                calculated_values_dict['Сумма продаж после разбивки, шт.:'] = new_values_dict_sum
                calculated_values_dict['Остаток после прогона №1:'] = remain
                iteration_count = 2

                while remain:
                    new_values_dict, remain = spread_remain(new_values_dict, remain)
                    calculated_values_dict[f'Остаток после прогона №{iteration_count}'] = remain
                    iteration_count += 1

                result_sum = sum([value for value in new_values_dict.values()])
                calculated_values_dict['Итоговая сумма значений после разбивки'] = result_sum

                result_values_list.append(new_values_dict)
                result_calculated_values.append(calculated_values_dict)

            # Обнуляем промежуточные словари и добавляем в него наименование и САП код текущего товара
            values_dict = {}
            calculated_values_dict = {}
            calculated_values_dict["SAP код"] = row[0].value
            calculated_values_dict["Наименование товара"] = row[1].value

        values_dict = create_dict_from_table(values_dict, row, start_column_values)

    return result_values_list, result_calculated_values


if __name__ == '__main__':
    app = App(main_func, create_result_file)
    app.mainloop()
